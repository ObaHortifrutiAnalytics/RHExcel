import sqlite3
from sqlite3 import Date
from pandas.io.sql import DatabaseError
import openpyxl
import paramiko
from paramiko import sftp_handle
from paramiko import sftp_client
import sqlalchemy

import pandas as pd
import cx_Oracle
# cx_Oracle.init_oracle_client(lib_dir=r"C:\oracle\instantclient_21_10",
#                              config_dir=r"C:\oracle\instantclient_21_10\network\admin")
import openpyxl
import config
import datetime
from dateutil.relativedelta import *
import smtplib 
import linecache
import sys
import timedelta
import os
from os import listdir, read
from os.path import isfile, join
from openpyxl import Workbook, load_workbook 
from openpyxl.styles import numbers
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )
from datetime import datetime, date
import time
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string,get_column_interval
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Color, numbers
from sqlalchemy.engine import create_engine
import smtplib, ssl
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from sqlalchemy.exc import DatabaseError
if os.name == "nt":
    os.system("cls")
else:
    os.system("clear")
# Variáveis
dPerInicial = ""
dPerFinal = ""
iInvalidos = 0
ListaInvalidos = []
ListRegisters = []
ListRegisters.append(['Tabela','Registros','Carregou','Erro', 'Query'])


global logFile

# Criar Lista e-Mails
global dEquipe
global dResponsavel
global dNivel1
global dNivel2
global dBackup
global dEmail
global dLoja
global dLojaResponsavel
global dLojaNivel1
global dLojaNivel2
global dLojaNivel3
global dBackupLoja
global dLojaEmail
global dCDs
global eMailDestino
global errEnvio
errEnvio = False
dEquipe = {}
dResponsavel = {}
dNivel1 = {}
dNivel2 = {}
dEmail = {}
dCDs = {}
dLoja = {}
dLojaResponsavel = {}
dLojaNivel1 = {}
dLojaNivel2 = {}
dLonaNivel3 = {}
dLojaEmail = {}
eMailDestino = ""
print("Preparando a Geração")
# Função
def fNoneBranco( sVariavel ):
    Resultado = ""
    if (sVariavel == None or sVariavel == "None"):
        Resultado = ""
    else:
        Resultado = str(sVariavel)
    return Resultado
def fnDataExec(sParam, bForce=False):
    if (config.dDataHoje == ''):
        dtDataDia = pd.to_datetime(datetime.now())  + pd.DateOffset(days=-1)
        #dtDataDia = datetime.now()
    else:
        dtDataDia = datetime.strptime(config.dDataHoje, '%Y_%m_%d')

    if (bForce==True):
        return dtDataDia.strftime('%Y_%m_%d')
    else:
        if (sParam[0:1] == 'd'): # Diária
            return dtDataDia.strftime('%Y_%m_%d')
        elif (sParam[0:1] == 'w'): # Semanal
            if (dtDataDia.weekday()==0):
                return dtDataDia.strftime('%Y_%m_%d')
            else: 
                return 'Não envie'
        elif (sParam[0:1] == 'm'): # Mensal
            if (dtDataDia.day == 1):
                return dtDataDia.strftime('%Y_%m_%d')
            else: 
                return 'Não envie'                
        elif (sParam[0:1] == 'b'): # Quinzenal
            if ((dtDataDia.day == 1) or (dtDataDia.day == 16)):
                return dtDataDia.strftime('%Y_%m_%d')
            else: 
                return 'Não envie'                
        elif (sParam[0:1] == 'y'): # Anual
            if ((dtDataDia.day == 1) and (dtDataDia.month == 1)): # Anual
                return dtDataDia.strftime('%Y_%m_%d')
            else: 
                return 'Não envie'                
        else: # Caso o parâmetro esteja incorreto
             return dtDataDia.strftime('%Y_%m_%d')
def fnCriarListaEmailsEquipe(dEquipe, dResponsavel, dNivel1, dNivel2, dEmail):
    wbNome = config.sDirEmails
    #Abrindo planilha
    wb = load_workbook(wbNome, read_only=False) 
    tsheet = wb['CDs']
    for col_cell in tsheet['A']:
        if (col_cell.row > 2): 
            CodFilial = fNoneBranco(tsheet.cell(row = col_cell.row, column = 9).value )
            CodFilial = fNoneBranco(CodFilial[0:CodFilial.find("_")])
            if (CodFilial != ""):
                dCDs[CodFilial] = CodFilial
    for col_cell in tsheet['A']:
        if (col_cell.row > 1):
            CodEquipe = fNoneBranco(col_cell.value)
            NomEquipe = fNoneBranco(tsheet.cell(row = col_cell.row, column = 2).value)
            Responsavel = fNoneBranco(tsheet.cell(row = col_cell.row, column = 3).value)
            Nivel1 = fNoneBranco(tsheet.cell(row = col_cell.row, column = 4).value)
            Nivel2 = fNoneBranco(tsheet.cell(row = col_cell.row, column = 5).value)
            sBackup = fNoneBranco(tsheet.cell(row = col_cell.row, column = 6).value)
            eMail = fNoneBranco(tsheet.cell(row = col_cell.row, column = 7).value)
            dEquipe[CodEquipe] = NomEquipe
            dResponsavel[CodEquipe] = Responsavel
            dNivel1[CodEquipe] = Nivel1
            dNivel2[CodEquipe] = Nivel2
            if (eMail == "1"):
                if (fNoneBranco(sBackup) != ""):
                    dEmail[CodEquipe] = Nivel1 + ";" + sBackup
                else:
                    dEmail[CodEquipe] = Nivel1    
            else:
                if (fNoneBranco(Nivel2) != ""):
                    if (fNoneBranco(sBackup) != ""):
                        dEmail[CodEquipe] = Nivel1 + ";" + sBackup + ";" + Nivel2 
                    else:   
                        dEmail[CodEquipe] = Nivel1 + ";" + Nivel2
                else:
                    if (fNoneBranco(sBackup) != ""):
                        dEmail[CodEquipe] = Nivel1 + ";" + sBackup
                    else:
                        dEmail[CodEquipe] = Nivel1
    wb.close
def fnCriarListaEmailsLoja(dLoja, dLojaResponsavel, dLojaNivel1, dLojaNivel2, dLojaEmail):
    wbNome = config.sDirEmails
    #Abrindo planilha
    wb = load_workbook(wbNome, read_only=False) 
    tsheet = wb['LOJAS']
    for col_cell in tsheet['A']:
        if (col_cell.row > 1) and (fNoneBranco(tsheet.cell(row = col_cell.row, column = 2).value) != ""):

            CodLoja = fNoneBranco(tsheet.cell(row = col_cell.row, column = 2).value)
            CodLoja = fNoneBranco(CodLoja[0:CodLoja.find("_")])

            LojaNomEquipe = fNoneBranco(tsheet.cell(row = col_cell.row, column = 2).value)
            LojaResponsavel = fNoneBranco(tsheet.cell(row = col_cell.row, column = 3).value)
            LojaNivel1 = fNoneBranco(tsheet.cell(row = col_cell.row, column = 4).value)
            LojaNivel2 = fNoneBranco(tsheet.cell(row = col_cell.row, column = 5).value)
            sBackupLoja = fNoneBranco(tsheet.cell(row = col_cell.row, column = 6).value)
            LojaeMail = fNoneBranco(tsheet.cell(row = col_cell.row, column = 7).value)
            dLoja[CodLoja] = LojaNomEquipe
            dLojaResponsavel[CodLoja] = LojaResponsavel
            dLojaNivel1[CodLoja] = LojaNivel1
            dLojaNivel2[CodLoja] = LojaNivel2
            # if (CodLoja == '27'):
            #     print(CodLoja)

            if (LojaeMail == "1"):
                if (sBackupLoja != ""):
                    dLojaEmail[CodLoja] = LojaNivel1 + ";" + sBackupLoja
                else:
                    dLojaEmail[CodLoja] = LojaNivel1    
            else:
                if (LojaNivel2 != ""):
                    if (sBackupLoja != ""):
                        dLojaEmail[CodLoja] = LojaNivel1 + ";" + sBackupLoja + ";" + LojaNivel2 
                    else:   
                        dLojaEmail[CodLoja] = LojaNivel1 + ";" + LojaNivel2
                else:
                    if (sBackupLoja != ""):
                        dLojaEmail[CodLoja] = LojaNivel1 + ";" + sBackupLoja
                    else:
                        dLojaEmail[CodLoja] = LojaNivel1
    wb.close
# Busca Destinatário e-Mail

def fnBuscaDestinatarioEmail():
    eMailDestino = []
    wbNome = config.sDirEmails
    #Abrindo planilha
    wb = load_workbook(wbNome, read_only=False) 
    tsheet = wb['EMAIL']
    for col_cell in tsheet['A']:
       sNome = fNoneBranco(tsheet.cell(row = col_cell.row, column = 1).value)
       if sNome != "":
            eMailDestino.append(sNome)
    return eMailDestino

# Grava linha do log
def fnLinhaLog(nFile, sText, sTipo = "Notificação"):
    sDataAgora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nFile.write(sDataAgora + ";" + sText + ";" + sTipo + '\n')
    nFile.close

# Refactoring
def fnBuscaPeriodo(iTipo):
    global dPerInicial, dPerFinal
    sDiaHoje = int(datetime.now().strftime('%d'))
    sMesHoje = int(datetime.now().strftime('%m'))
    sAnoHoje = int(datetime.now().strftime('%y'))
    if (sDiaHoje <= 17):
        if (sMesHoje == 1):
            sMesHoje = 12
            sAnoHoje = sAnoHoje - 1
        else:
            sMesHoje = sMesHoje - 1
    sDiaHoje = 16
    # dPerInicial = sDiaHoje + '/' + sMesHoje + '/' + sAnoHoje
    dPerInicial = date(int(sAnoHoje)+2000, int(sMesHoje), sDiaHoje)
    dPerInicial = dPerInicial.strftime('%d/%m/%y')

    dPerFinal = date(int(sAnoHoje)+2000,int(sMesHoje), 15) 
    dPerFinal = dPerFinal + relativedelta(months=+1) 


    if (dPerFinal >= date(int(sAnoHoje)+2000, int(sMesHoje), sDiaHoje)):
        if eMailDestino == "Adm":
            dPerFinal = datetime.now() + relativedelta(days=-1)    
        else:
            if sDiaHoje == 17:
                dPerFinal = datetime.now() + relativedelta(days=-1)  
            else :
                dPerFinal = datetime.now() + relativedelta(days=-2)    

    dPerFinal = dPerFinal.strftime('%d/%m/%y')
    

def fnEnviaEmails(iTipo, sDestinatario, wbNome, ExcelFileName, logFile):
    if config.bSendEmail:
        message = MIMEMultipart("alternative")
        if config.sEmailTeste != "":
            message["Subject"] = config.sMailSubject + " - Teste"
        else:    
            message["Subject"] = config.sMailSubject
        message["From"] = config.sSMTPUserName
        # if config.sEmailTeste != "":
        #     message["To"] = config.sEmailTeste
        # else:
        message["To"] = sDestinatario
        if iTipo == 1: # Mensagem para Administração
            # Mensagem 1
            text = """\
            Seguem os apontamentos de horas realizados para todas as áreas. Favor corrigir e/ou tomar as ações necessárias. """
            if not errEnvio:
                html = """\
                <html>
                <body>
                    <p>Olá, tudo bem??<br>
                    Em anexo segue planilha Apuração de Ponto com as divergências e pontos de atenção para validação e tratativas, conforme aplicável. <br><br>
                    Os dados foram gerados considerando o período inical: <b>{}</b> até o período final <b>{}</b><br><br>
                    Qualquer dúvida favor abrir chamado no <b><a href='https://fluig.redeoba.com.br/portal/p/01/pageprocessstart'>Service Desk</a></b>.
                    <br>
                    Em Central de Tarefas => BACKOFFICE (Ponto/Admissão/Uniforme) => PP (Selecione conforme aplicável)
                    <br>
                    <b>Por favor, não responda esse e-mail. Não haverá retorno</b>
                    <br><br>
                    Atenciosamente<br>
                    Time de Ponto - Gente e Gestão
                    <br>
                    <H1><FONT COLOR="GREEN">E-mails enviados aos usuários.</FONT></H1>
                    </p>
                </body>
                </html>
                """.format(dPerInicial, dPerFinal)
            else:
                html = """\
                <html>
                <body>
                    <p>Olá, tudo bem??<br>
                    Em anexo segue planilha Apuração de Ponto com as divergências e pontos de atenção para validação e tratativas, conforme aplicável. <br><br>
                    Os dados foram gerados considerando o período inical: <b>{}</b> até o período final <b>{}</b><br><br>
                    Qualquer dúvida favor abrir chamado no <b><a href='https://fluig.redeoba.com.br/portal/p/01/pageprocessstart'>Service Desk</a></b>.
                    <br>
                    Em Central de Tarefas => BACKOFFICE (Ponto/Admissão/Uniforme) => PP (Selecione conforme aplicável)
                    <br>
                    <b>Por favor, não responda esse e-mail. Não haverá retorno</b>
                    <br><br>
                    Atenciosamente<br>
                    Time de Ponto - Gente e Gestão
                    <br>
                    <H1><FONT COLOR="RED">O e-mail não será enviado por erro. Favor olhar o log gerado!</FONT></H1>
                    </p>
                </body>
                </html>
                """.format(dPerInicial, dPerFinal)
        elif iTipo ==  2: # Mensagem para Gestor
            # Mensagem 1
            text = """\
            Seguem os apontamentos da sua área. Favor corrigir e/ou tomar as ações necessárias. """
            html = """\
            <html>
            <body>
                    <p>Olá, tudo bem??<br>
                    Em anexo segue planilha Apuração de Ponto com as divergências e pontos de atenção para validação e tratativas, conforme aplicável. <br><br>
                    Os dados foram gerados considerando o período inical: <b>{}</b> até o período final <b>{}</b><br><br>
                    Qualquer dúvida favor abrir chamado no <b><a href='https://fluig.redeoba.com.br/portal/p/01/pageprocessstart'>Service Desk</a></b>.
                    <br>
                    Em Central de Tarefas => BACKOFFICE (Ponto/Admissão/Uniforme) => PP (Selecione conforme aplicável)
                    <br>
                    <b>Por favor, não responda esse e-mail. Não haverá retorno</b>
                    <br><br>
                    Atenciosamente<br>
                    Time de Ponto - Gente e Gestão
                    <br>
                    </p>
            </body>
            </html>
            """.format(dPerInicial, dPerFinal)
        
        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")

        message.attach(part1)
        message.attach(part2)

        if iTipo == 1: # Anexo para Administração
            # Anexando a Planilha
            msg_attach = email.mime.base.MIMEBase('application', 'octet-stream')
            with open(wbNome, 'rb') as f:
                    msg_attach.set_payload(f.read())

            email.encoders.encode_base64(msg_attach)
            msg_attach.add_header('Content-Disposition','attachment',  filename= ExcelFileName )        

            message.attach(msg_attach)
            # Anexando o Log
            msg_attach = email.mime.base.MIMEBase('application', 'text/plain')
            logFile.close
            with open(config.sLogDirectory + nLogFile, 'rb') as f:
                    msg_attach.set_payload(f.read())

            email.encoders.encode_base64(msg_attach)
            msg_attach.add_header('Content-Disposition','attachment', filename= nLogFile )        
            logFile = open(config.sLogDirectory + nLogFile, 'a')
            message.attach(msg_attach)
        elif iTipo == 2: # Anexo Administrador
            # Anexando a Planilha
            msg_attach = email.mime.base.MIMEBase('application', 'octet-stream')
            with open(wbNome, 'rb') as f:
                    msg_attach.set_payload(f.read())
            email.encoders.encode_base64(msg_attach)
            msg_attach.add_header('Content-Disposition','attachment',  filename= ExcelFileName )        
            message.attach(msg_attach)
        context = ssl.create_default_context()
        server = smtplib.SMTP(config.sSMTP, config.iSMTPport)
        server.starttls()
        server.login(config.sSMTPUserName, config.sSMTPUserPWD)
        if config.sEmailTeste != "":
            sDestinatario = config.sEmailTeste
        server.sendmail(config.sSMTPUserName, sDestinatario, message.as_string())
        server.quit()
    else:
        fnLinhaLog(logFile, 'Configurado para não enviar e-mail em config.bSendEmail ao destinatário: ' + sDestinatario)

def FormataPlanilha(sPlanilha, sNomeDestino, iDestino):
    # Formatando Excel
    fnLinhaLog(logFile, 'Formatando a planilha: ' + sNomeDestino + " " + str(datetime.now().strftime('%d/%m/%Y %H:%M:%S')))
    #Abrindo planilha
    wb = load_workbook(sPlanilha, read_only=False)  
    # print("Tratando campos")
    tsheets = wb.sheetnames
    tsheet = wb.active
    tsheet.title= "Base de Dados"
    # print(tsheet.title)


    # Formatar Data
    # print("Formatando data")
    dia = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom']
    for col_cell in tsheet['J']:
        col_cell.number_format =  numbers.FORMAT_DATE_DDMMYY
        if (col_cell.row > 1):
            colRow = 'K' + str(col_cell.row)
            tsheet[colRow] = dia[col_cell.value.weekday()]




    # Colorir Colunas
    my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    for col_cell in tsheet['A']:
        if (col_cell.row == 1):
                tsheet.cell(row = col_cell.row, column = 2).fill = my_fill
                tsheet.cell(row = col_cell.row, column = 2).font = Font(color="FFFFFF", italic=False, bold=True)
                tsheet.cell(row = col_cell.row, column = 13).fill = my_fill
                tsheet.cell(row = col_cell.row, column = 13).font = Font(color="FFFFFF", italic=False, bold=True)
                tsheet.cell(row = col_cell.row, column = 22).fill = my_fill
                tsheet.cell(row = col_cell.row, column = 22).font = Font(color="FFFFFF", italic=False, bold=True)



    # Ocultar Colunas
    tsheet.column_dimensions['A'].hidden = True
    tsheet.column_dimensions['F'].hidden = True
    tsheet.column_dimensions['W'].hidden = True
    tsheet.column_dimensions['Y'].hidden = True
    tsheet.column_dimensions['Z'].hidden = True
    tsheet.column_dimensions['AA'].hidden = True
    tsheet.column_dimensions['AB'].hidden = True


    # Formatar Tipo
    for col_cell in tsheet['H']:
        col_cell.number_format =  '@'


    # Incluir filtr5os
    tsheet.auto_filter.ref = tsheet.dimensions

    # Excluir Colunas
    if iDestino == 2:
        tsheet.column_dimensions['A'].hidden = False
        tsheet.column_dimensions['F'].hidden = False
        tsheet.column_dimensions['W'].hidden = False
        tsheet.column_dimensions['Y'].hidden = False
        tsheet.column_dimensions['Z'].hidden = False
        tsheet.column_dimensions['AA'].hidden = False
        tsheet.column_dimensions['AB'].hidden = False
        tsheet.delete_cols(29)
        tsheet.delete_cols(28)
        tsheet.delete_cols(27)
        tsheet.delete_cols(26)
        tsheet.delete_cols(25)
        tsheet.delete_cols(23)
        tsheet.delete_cols(6)
        tsheet.delete_cols(5)
        tsheet.delete_cols(3)
        tsheet.delete_cols(1)    

    # Resize

    dims = {}
    for row in tsheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for col, value in dims.items():
        tsheet.column_dimensions[col].width = value + 10

    # Tirar Grid
    tsheet.sheet_view.showGridLines = False

    # Salvar Planilha
    wb.save(sPlanilha)
    # Separar Pastas


#
# Hora de Início
#
if (config.dDataHoje == ''):
    data_atual = pd.to_datetime(datetime.now())  + pd.DateOffset(days=-1)
    data_atual = data_atual.strftime('%Y_%m_%d') 
else:
    data_atual = config.dDataHoje

data_inicio = datetime.now().strftime('%d-%m-%Y %H:%M')


# Abrir o arquivo de Log
nLogFile =   data_atual + "_" + config.sLogFile + "_" + datetime.now().strftime('%H_%M_%S') + '.log'
logFile = open(config.sLogDirectory + nLogFile, 'w')

fnLinhaLog(logFile, "Início da Preparação")
# Lendo parâmetros
if __name__ == "__main__":
    iTipo = int(os.sys.argv[1])
    #iTipo = 2
    if iTipo == 1:
        eMailDestino = "Adm"
    else:
        eMailDestino = "Equipe"
if (iTipo == 1):
    fnLinhaLog(logFile,"Parâmetro de Execução Administração " + str(iTipo), "Alerta")
else:
    fnLinhaLog(logFile,"Parâmetro de Execução Gestor " + str(iTipo), "Alerta")   

try:
#
# Apagar Destino
#
    fnLinhaLog(logFile, "Apagando arquivos do envio anterior")
    onlyfiles = [f for f in listdir(config.sDiretorioOrigem) if isfile(join(config.sDiretorioOrigem, f))]
    for filename in onlyfiles:
        os.remove(config.sDiretorioOrigem + "\\" + filename )
    
    fnLinhaLog(logFile, "Arquivos do dia anterior excluídos")
    # Apagando os logs antigos
    # Usa como base o config.iPerRetencao
    onlyfiles = [f for f in listdir(config.sLogDirectory) if isfile(join(config.sLogDirectory, f))]
    DiasRetencao = timedelta.Timedelta(days=config.iPerRetencao+1)
    dtRetencao = datetime.now() - DiasRetencao
    for filename in onlyfiles:
        if (filename[-4:] == ".log"):
            # print(filename[0:10], dtRetencao.strftime('%Y_%m_%d'))
            sData = filename[0:10]
            sData.replace("-","_")
            sData.replace("/","_")
            try:
                data_file = datetime.strptime(sData, '%Y_%m_%d')
            except:
                fnLinhaLog(logFile, "Formato de data inválido: " + filename)
            if (data_file < dtRetencao):        
                os.remove(config.sLogDirectory + "\\" + filename )
                fnLinhaLog(logFile,"Apaguei " + filename)
    # Abrir banco Oracle
    fnLinhaLog(logFile, "Abrindo conexão com o banco de dados")

    print("abrindo banco de dados oracle")
    fnLinhaLog(logFile, "Abrindo conexão com o banco de dados")
    #
    # Telemetria
    #
    diretorio = os.getcwd()
    cursor = None
    r_codevento = None
    DIALECT = config.dialet
    SQL_DRIVER = config.sql_driver
    USERNAME = config.usernameTelemetria #enter your username
    PASSWORD = config.passwordTelemetria #enter your password
    HOST = config.hostTelemetria #enter the oracle db host url
    PORT = config.port # enter the oracle port number
    SERVICE = config.serviceTelemetria # enter the oracle db service name
    ENGINE_PATH_WIN_AUTH = DIALECT + '+' + SQL_DRIVER + '://' + USERNAME + ':' + PASSWORD +'@' + HOST + ':' + str(PORT) + '/?service_name=' + SERVICE
    if config.bTelemetria:    
        connTelemetria = cx_Oracle.connect(
                    config.usernameTelemetria,
                    config.passwordTelemetria,
                    config.dsnTelemetria,
                    encoding=config.encoding) 
        cursor = connTelemetria.cursor()
        r_codevento = cursor.var(int)
        fnLinhaLog(logFile, "Conexão com o banco de dados Telemetria realizada")        
        #Entrada Telemetria
        cursor.callproc('telemetria.pkg_ger_eventos.p_criaevento', [config.codEntradaTelemetria,config.sObjetivoTelemetria,diretorio,config.username,r_codevento])
        fnLinhaLog(logFile, "Entrada de evento telemetria gerado" + str(r_codevento))
    #    
    # Transação
    #
    conn = cx_Oracle.connect(
                config.username,
                config.password,
                config.dsn,
                encoding=config.encoding) 
    fnLinhaLog(logFile, "Conexão com o banco de dados realizada")

        
    fnBuscaPeriodo(iTipo) # dPerInicial e dPerFinal
    fnLinhaLog(logFile, 'Periodo Inicial:' + dPerInicial)
    fnLinhaLog(logFile, 'Periodo Final:' + dPerFinal)
    # Bloquear após o dia 17 por pelo período definido em config.iDiasBloqueio
    sDiaHoje = int(datetime.now().strftime('%d'))
    bBloqueioEnvio = False
    if sDiaHoje > 17:
        dataParte = str(dPerFinal)
        iDay = 17
        iMonth = int(datetime.now().strftime('%m'))
        iYear = int(datetime.now().strftime('%Y'))
        dFinal = date(iYear, iMonth, iDay)
        iDiaFinalBloqueio =  dFinal  + relativedelta(days=+config.iDiasBloqueio)
        if iDiaFinalBloqueio.strftime('%y/%m/%d') >= datetime.now().strftime('%y/%m/%d'):
            bBloqueioEnvio = True
    if not bBloqueioEnvio or config.bForce:
        fnCriarListaEmailsEquipe(dEquipe, dResponsavel, dNivel1, dNivel2, dEmail)
        fnCriarListaEmailsLoja(dLoja, dLojaResponsavel, dLojaNivel1, dLojaNivel2, dLojaEmail)
        sHeaders = ["Problema","Tipo de Ocorrência", "Cod Filial", "Filial", "CodEquipe","Seção","Equipe","Chapa","Funcionário","Data","Dia Semana","Observação","Horas Ocorrência","Bat1", "Bat2", "Bat3", "Bat4", "Bat5", "Bat6", "Bat7", "Bat8", "Problema Detalhado", "Banco de Horas", "Horário Normal", "Query", "Equipe e-Mail", "Responsavel", "e-Mail", "Erro"]
        # Reduzir um número - Começa com zero
        sHeadersCol = [
            # [1, 3, 2, 4, 6, 7, 8, 9, 10, 25, 21, -1, -1, 61, 2, 3, 4, 5, 6, 7, 8], # RM.PLN_PONTO_REFEICAO
            [16, 19, 18, 20, 3, 4, 0, 1, 2, 24, 5, -1, -1,  6, 7, 8, 9, 10, 11, 12, 13], # RM.PLN_PONTO_IMPAR
            [1, 3, 2, 4, 6, 7, 8, 9, 10, 25, 21, -1, -1, 12, 13, 14, 15, 16, 17, 18, 19], # RM.PLN_PONTO_FALTA
            [1, 3, 2, 4, 6, 7, 8, 9, 10, 27, 21, -1, -1, 12, 13, 14, 15, 16, 17, 18, 19], # RM.PLN_PONTO_ATRASO
            [1, 3, 2, 4, 6, 7, 8, 9, 10, 30, 23, -1, -1, 12, 13, 14, 15, 16, 17, 18, 19], # RM.PLN_PONTO_EXTRAS
            # [0, 7, 6, 8, 3, 4, -1, -1, -1, 27, -1, 28], # RM.PLN_SALDOBANHORAS
            [2, 4, 8, 5, 6, 7, 12, -1, -1, 17, 9, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1], # RM.PLN_OCOR_INTERJORNADAS -- Não tem batida
            [2, 4, 8, 5, 6, 7, 9, -1, -1, 17, 10, -1, -1, 11, 12, -1, -1, -1, -1, -1, -1], # RM.PLN_OCOR_REFEICAO -- Não tem batida
            [2, 4, 9, 5, 6, 7, 11, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1], # RM.PLN_OCOR_7DIASCONSEC -- Não tem batida
            [2, 4, 9, 5, 7, 8, 11, -1, -1, 16, -1, -1, -1,  -1, -1, -1, -1, -1, -1, -1, -1],  # RM.PLN_3DOM_CONSECUTIVOS -- Não tem batida
            # [6, 11, 10, 1, 2, 8, 15, 16, 18, 14, 22, -1, 31],  # RM.PLN_OCOR_ABONGERALFIL
            [6, 11, 10, 12, 1, 2, 15, 16, 18, 14, 22, -1, 31,  -1, -1, -1, -1, -1, -1, -1, -1], # RM.PLN_OCOR_ABONGERAL -- Não tem batida
            [1, 3, 2, 4, 6, 7, 8, 9, 10, 30, 23, -1, -1, 12, 13, 14, 15, 16, 17, 18, 19]  # RM.PLN_OCOR_HEXTRAS
        ]
        sQueries = [
            # --"SELECT * FROM RM.PLN_PONTO_REFEICAO  where data >= '{}' and data <= '{}';".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_PONTO_IMPAR  where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_PONTO_FALTA where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_PONTO_ATRASO where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_PONTO_EXTRAS where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            #-- "SELECT * FROM RM.PLN_SALDOBANHORAS;",
            "SELECT * FROM RM.PLN_OCOR_INTERJORNADAS where databatida between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_OCOR_REFEICAO where databatida between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_OCOR_7DIASCONSEC where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_3DOM_CONSECUTIVOS where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            #-- "SELECT * FROM RM.PLN_OCOR_ABONGERALFIL where data >= '{}' and data <= '{}';".format(dPerInicial, dPerFinal),        
            "SELECT * FROM RM.PLN_OCOR_ABONGERAL where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            "SELECT * FROM RM.PLN_OCOR_HEXTRAS where data between to_date('{}', 'dd/mm/yy') and to_date('{}', 'dd/mm/yy');".format(dPerInicial, dPerFinal),
            ]
        sProblemas = [
            
            # "REFEIÇÃO: O intervalo mínimo de 1 hora de almoço, de acordo com a CLT, é concedido para quem segue uma jornada de trabalho acima de 6 horas.",
            "Marcações impares (falta de marcação)",
            "Validar faltas",
            "Validar o atraso",
            "Horas Extras executadas",        
            # "Banco de Horas",
            "INTERJORNADA: Descanso entre duas jornadas de trabalho de, no mínimo, 11 horas consecutivas.",        
            "REFEIÇÃO: O intervalo mínimo de 1 hora de almoço, de acordo com a CLT, é concedido para quem segue uma jornada de trabalho acima de 6 horas.",    
            "7º JORNADA: O colaborador que trabalha durante 6 dias consecutivos, é necessário folgar no 7º dia",
            "3º DOMINGO: De acordo com a legislação trabalhista o colaborador pode trabalhar, no máximo dois domingos seguidos.",   
            # "ABONO FILIAL - Demonstra a quantidade de horas abonadas separadas por filial e por tipo de abono e o valor que representa para a companhia sobre cada hora abonada a ação necessária é validar o motivo de tantos abonos sendo realizados e formas de mitigar pela Gestor",        
            "ABONO COMPLETO - Quantidade e valor de horas/dias abonados, por tipo de abono",
            "+2 HORAS EXTRAS: De acordo com a legislação não é permitido exceder mais de 2 horas extras diárias."
            ]
        sProblemasReduzidos = [
            # "INCLUIR REFEIÇÃO",
            "ADVERTIR",
            "FALTAS",
            "ATRASOS",
            "+2 HORAS EXTRAS",        
            # "Banco de Horas",
            "INTERJORNADA",        
            "REFEIÇÃO",    
            "7º JORNADA",
            "3º DOMINGO",   
            # "ABONO FILIAL",        
            "ABONO COMPLETO",
            "+2HE"
            ]
        sTipo = [
            # "Validar Informação",
            "Inconsistência",
            "Validar Informação",
            "Validar Informação",
            "Validar Informação",        
            # "SELECT * FROM RM.PLN_SALDOBANHORAS;",
            "Infração",    
            "Infração",  
            "Infração",
            "Infração", 
            # "Validar Informação",      
            "Validar Informação",
            "Infração"
            ]
        print('Início: ', datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        #Agrupar Dados
        dataDestino = []
        for i in range(len(sQueries)):
            #dataDestino.append(sHeaders)
            query = sQueries[i]
            query = query.replace(";", "") # Tirar o ; da query
            # print(query)
            dataframe = pd.read_sql(query, conn)    
            index = dataframe.index

            sTexto = "Query: " + query + " Registros: " + fNoneBranco(len(index))
            print(sTexto)
            fnLinhaLog(logFile, sTexto)
            if (len(index) != 0): 
                for iContador in range(len(index)):
                    sErro = ""
                    if sHeadersCol[i][0] == -1:
                        Filial = ""
                        CodFilial = ""
                    else:
                        # Filial[0:Filial.find("_")]
                        if (dataframe.iat[iContador,sHeadersCol[i][0]]=="FRESHLABS"):
                            codFilial = "FRESHLABS"
                            Filial = "FRESHLABS"
                        else:
                            CodFilial = dataframe.iat[iContador,sHeadersCol[i][0]]
                            CodFilial = fNoneBranco(CodFilial[0:CodFilial.find("_")])
                            Filial = dataframe.iat[iContador,sHeadersCol[i][0]]
                            Filial = Filial[Filial.find("_")+1:]
                    if sHeadersCol[i][1] == -1:
                        CodEquipe = ""
                    else:
                        CodEquipe = fNoneBranco(dataframe.iat[iContador,sHeadersCol[i][1]])
                    if sHeadersCol[i][2] == -1:
                        Secao = ''
                    else:
                        Secao = dataframe.iat[iContador,sHeadersCol[i][2]]
                    if sHeadersCol[i][3] == -1:
                        Equipe = ''
                    else:
                        Equipe = dataframe.iat[iContador,sHeadersCol[i][3]]                   
                    if sHeadersCol[i][4] == -1:
                        Chapa = ''
                    else:
                        Chapa = dataframe.iat[iContador,sHeadersCol[i][4]]                
                    if sHeadersCol[i][5] == -1:
                        Funcionario = ''
                    else:
                        Funcionario = dataframe.iat[iContador,sHeadersCol[i][5]]
                    if sHeadersCol[i][6] == -1:
                        Data = ''
                    else:
                        Data = dataframe.iat[iContador,sHeadersCol[i][6]]
                    if sHeadersCol[i][7] == -1:
                        DiaSemana = ''
                    else:
                        DiaSemana = dataframe.iat[iContador,sHeadersCol[i][7]]
                    if sHeadersCol[i][8] == -1:
                        Observacao = ''
                    else:
                        Observacao = dataframe.iat[iContador,sHeadersCol[i][8]]
                    if sHeadersCol[i][9] == -1:
                        HorarioNormal = ''
                    else:
                        HorarioNormal = dataframe.iat[iContador,sHeadersCol[i][9]]
                    if sHeadersCol[i][10] == -1:
                        HorasOcorrencia = ''
                    else:
                        HorasOcorrencia = str(dataframe.iat[iContador,sHeadersCol[i][10]]).replace(" ","")
                        # HorasOcorrencia = ''.join(char for char in HorasOcorrencia if char.isalnum())

                    if sHeadersCol[i][11] == -1:
                        Problema = sProblemas[i]
                    else:
                        Problema = "Não identificado"
                    if sHeadersCol[i][12] == -1:
                        BancodeHoras = ""
                    else:
                        BancodeHoras = dataframe.iat[iContador,sHeadersCol[i][11]]
                    if sHeadersCol[i][13] == -1:
                        Bat1 = ''
                    else:
                        Bat1 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][13]]).replace(" ",""))
                        if ("+" in Bat1) or ("®" in Bat1):
                            if sErro == "":
                                sErro = "Bat1 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat1 tem deslocamento"                            
                    if sHeadersCol[i][14] == -1:
                        Bat2 = ''
                    else:
                        Bat2 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][14]]).replace(" ",""))
                        if ("+" in Bat2) or ("®" in Bat2):
                            if sErro == "":
                                sErro = "Bat2 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat2 tem deslocamento"      
                    if sHeadersCol[i][15] == -1:
                        Bat3 = ''
                    else:
                        Bat3 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][15]]).replace(" ",""))          
                        if ("+" in Bat3) or ("®" in Bat3):
                            if sErro == "":
                                sErro = "Bat3 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat3 tem deslocamento"                
                    if sHeadersCol[i][16] == -1:
                        Bat4 = ''
                    else:
                        Bat4 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][16]]).replace(" ",""))    
                        if ("+" in Bat4) or ("®" in Bat4):
                            if sErro == "":
                                sErro = "Bat4 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat4 tem deslocamento"                    
                    if sHeadersCol[i][17] == -1:
                        Bat5 = ''
                    else:
                        Bat5 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][17]]).replace(" ",""))   
                        if ("+" in Bat5) or ("®" in Bat5):
                            if sErro == "":
                                sErro = "Bat5 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat5 tem deslocamento"                     
                    if sHeadersCol[i][18] == -1:
                        Bat6 = ''
                    else:
                        Bat6 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][18]]).replace(" ",""))   
                        if ("+" in Bat6) or ("®" in Bat6):
                            if sErro == "":
                                sErro = "Bat6 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat6 tem deslocamento"                        
                    if sHeadersCol[i][19] == -1:
                        Bat7 = ''
                    else:
                        Bat7 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][19]]).replace(" ",""))    
                        if ("+" in Bat7) or ("®" in Bat7):
                            if sErro == "":
                                sErro = "Bat7 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat7 tem deslocamento"                       
                    if sHeadersCol[i][20] == -1:
                        Bat8 = ''
                    else:
                        Bat8 = fNoneBranco(str(dataframe.iat[iContador,sHeadersCol[i][20]]).replace(" ",""))
                        if ("+" in Bat8) or ("®" in Bat8):
                            if sErro == "":
                                sErro = "Bat8 tem deslocamento"
                            elif sErro != "":
                                sErro = sErro + "; Bat8 tem deslocamento"       
                    if (CodEquipe != ""):
                        if CodFilial in dCDs:
                            if CodEquipe in dEquipe:
                                EquipeMail = dEquipe[CodEquipe] 
                                Responsavel = dResponsavel[CodEquipe]
                                eMail = dEmail[CodEquipe]  
                            else:
                                sErro = "A equipe " + CodEquipe + " não existe na lista de e-mails "
                                fnLinhaLog(logFile, sErro, "Erro")
                                EquipeMail = ""
                                Responsavel = ""
                                eMail = ""  
                                errEnvio = True
                        else:
                            if CodFilial in dLoja:
                                # if (CodFilial == '27'):
                                #     print(CodFilial)
                                EquipeMail = dLojaEmail[CodFilial] 
                                Responsavel = dLojaResponsavel[CodFilial]
                                eMail =  dLojaEmail[CodFilial]
                            else:
                                fnLinhaLog(logFile, "A filial " + CodFilial + " não existe na lista de e-mails ", "Erro")     
                                EquipeMail = ""
                                Responsavel = ""
                                eMail = ""    
                                errEnvio = True                                                                         
                    dataDestino.append([sProblemasReduzidos[i], sTipo[i], CodFilial, Filial, CodEquipe,Secao,Equipe,Chapa,Funcionario,Data,DiaSemana,Observacao,HorasOcorrencia,Bat1, Bat2, Bat3, Bat4, Bat5, Bat6, Bat7, Bat8, Problema, BancodeHoras, HorarioNormal, sQueries[i], EquipeMail, Responsavel, eMail, sErro])
            else:
                fnLinhaLog(logFile, "Arquivo sem linhas :" + query, "Erro")
        dataGeral = pd.DataFrame(dataDestino, columns=sHeaders)
        blankIndex=[''] * len(dataGeral)
        dataGeral.index=blankIndex
        dfDestino = dataGeral.sort_values(['Equipe','Filial','Funcionário'])
        ExcelFileName = data_atual + '_' + config.sNomePlanilha
        dfDestino.to_excel(config.sDiretorioOrigem + ExcelFileName, index=False)
        fnLinhaLog(logFile, 'Término Obter Dados: ' + str(datetime.now().strftime('%d/%m/%Y %H:%M:%S')))
        # Formatando Excel
        fnLinhaLog(logFile, 'Início Formatar Excel: ' + str(datetime.now().strftime('%d/%m/%Y %H:%M:%S')))
        wbNome = config.sDiretorioOrigem + data_atual + "_" + config.sNomePlanilha
        FormataPlanilha(wbNome, "Administração RH", 1)  # Força o iTipo para ir completa
        # Convertendo e-mails (equipes)
        fnLinhaLog(logFile, 'Gravando lista de e-Mails')
        wb = Workbook()
        sheet = wb.active
        sheet["A1"] = "Equipe"
        sheet["B1"] = "eMail"
        for row, (Equipe, Email) in enumerate(dEmail.items(), start=1):
            sEmail = Email.replace(";", ",")
            sheet [f"A{row}"] = Equipe
            sheet [f"B{row}"] = sEmail
        wb.save(config.sDiretorioOrigem + "EmailEquipe.xlsx")
        wb.close
        wb = Workbook()
        sheet = wb.active
        # Convertendo e-mails (empresas)
        sheet["A1"] = "Empresa"
        sheet["B1"] = "eMail"
        for row, (Loja, Email) in enumerate(dLojaEmail.items(), start=1):
            sEmail = Email.replace(";",",")
            sheet [f"A{row}"] = Loja
            sheet [f"B{row}"] = sEmail
        wb.save(config.sDiretorioOrigem + "EmailEmpresa.xlsx")
        wb.close
        fnLinhaLog(logFile, 'Concluindo lista de e-Mails')
        fnLinhaLog(logFile, 'Término Formatar Excel: ' + str(datetime.now().strftime('%d/%m/%Y %H:%M:%S')))
        # Enviar e-mail
        fnLinhaLog(logFile, "Enviando e-mails aos usuários","Alerta" )
        fnLinhaLog(logFile, 'Preparando o e-mail')
        Destinatario = fnBuscaDestinatarioEmail()
        for sDestinatario in Destinatario:
            fnEnviaEmails(iTipo, sDestinatario,wbNome, ExcelFileName, logFile )
        if (iTipo == 2 and not errEnvio):
            wb = load_workbook(wbNome, read_only=True)  
            ExcelGeral = pd.read_excel(wbNome)
            # Enviando e-mails de Equipe
            wbEquipe = load_workbook(config.sDiretorioOrigem + "EmailEquipe.xlsx", read_only=False)
            tsheets = wbEquipe.sheetnames
            tsheet = wbEquipe.active
            for col_cell in tsheet['A']:
                if fNoneBranco(tsheet.cell(row = col_cell.row, column = 1).value) != "":
                    sEquipe = float(tsheet.cell(row = col_cell.row, column = 1).value)
                    sEmail = str(tsheet.cell(row = col_cell.row, column = 2).value)
                    sEmail =  sEmail.replace(";",",")
                    sEmail = sEmail.split(",")
                    df = ExcelGeral[ExcelGeral['CodEquipe']==sEquipe]
                    df.to_excel(config.sDiretorioOrigem + "\\" + config.sNomePlanilhaEquipe, index=False)   
                    FormataPlanilha(config.sDiretorioOrigem + "\\" + config.sNomePlanilhaEquipe, "Equipe: " + str(sEquipe), 2)
                    Destinatario = sEmail
                    for sDestinatario in Destinatario:
                        if sDestinatario.find("@") > 0:
                            fnEnviaEmails(iTipo, sDestinatario,config.sDiretorioOrigem + "\\" + config.sNomePlanilhaEquipe, config.sNomePlanilhaEquipe, logFile )
                            fnLinhaLog(logFile,'Enviado e-mail para Equipe:' + str(sEquipe) + " eMail: " + sDestinatario)  
                        else:
                            fnLinhaLog(logFile,'Email não está correto: Equipe ' + str(sEquipe) + " eMail: " + sDestinatario)
            wbEmpresa = load_workbook(config.sDiretorioOrigem + "EmailEmpresa.xlsx", read_only=False)   
            tsheets = wbEmpresa.sheetnames
            tsheet = wbEmpresa.active
            for col_cell in tsheet['A']:
                if fNoneBranco(tsheet.cell(row = col_cell.row, column = 1).value) != "":
                    sEquipe = tsheet.cell(row = col_cell.row, column = 1).value
                    sEmail = tsheet.cell(row = col_cell.row, column = 2).value
                    sEmail = sEmail.replace(";",",")
                    sEmail = sEmail.split(",")
                    if (sEquipe == "121"):
                        df = ExcelGeral[(ExcelGeral['Cod Filial']==float(sEquipe)) & (ExcelGeral['Filial']!='FRESHLABS')]
                    else:
                        df = ExcelGeral[ExcelGeral['Cod Filial']==float(sEquipe)]
                    df.to_excel(config.sDiretorioOrigem + "\\" + config.sNomePlanilhaEquipe, index=False)   
                    FormataPlanilha(config.sDiretorioOrigem + "\\" + config.sNomePlanilhaEquipe, "Empresa: " + str(sEquipe), 2)
                    Destinatario = sEmail
                    for sDestinatario in Destinatario:
                        if sDestinatario.find("@") > 0:
                            fnEnviaEmails(iTipo, sDestinatario,config.sDiretorioOrigem + "\\" + config.sNomePlanilhaEquipe, config.sNomePlanilhaEquipe, logFile )
                            fnLinhaLog(logFile,'Enviado e-mail para Empresa:' + str(sEquipe) + " eMail: " + sDestinatario)  
                        else:
                            fnLinhaLog(logFile,'Email não está correto: Empresa ' + str(sEquipe) + " eMail: " + sDestinatario)
            if config.bForce:
                fnLinhaLog(logFile,'Força bruta atividada, nessa condição gera o arquivo sempre')    
    else:
        fnLinhaLog(logFile,'Periodo de Freezing até o dia : ' + iDiaFinalBloqueio.strftime('%d/%m/%Y'))
        fnLinhaLog(logFile,'Não gera arquivo nem envia e-mails em ' + str(config.iDiasBloqueio) + " dias")
except:
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    lineno = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, lineno, f.f_globals)
    sTexto = 'EXCEPTION IN ({}, LINE {} "{}"): {}'.format(filename, lineno, line.strip(), exc_obj)
    fnLinhaLog(logFile,"!!!!!!!!!!!!!!!!!! Ocorreu um erro : " + sTexto, "Erro")
    if config.bTelemetria:
        cursor.callproc('telemetria.pkg_ger_eventos.p_atualizaevento', [sTexto,r_codevento])
        cursor.close()
        connTelemetria.close()
    conn.close()
    logFile.close
    print("Acabou com erro: " + sTexto)
finally:
    fnLinhaLog(logFile, 'Acabou ********************************')
    print('Término: ', datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    print('Período Inicial: ' + dPerInicial)
    print('Período Final: ' + dPerFinal)
    print("\n\n***************************** Acabou **************************************************")
    if config.bTelemetria:
        cursor.callproc('telemetria.pkg_ger_eventos.p_atualizaevento', [None,r_codevento])
        cursor.close()
        connTelemetria.close()
    conn.close()

    logFile.close